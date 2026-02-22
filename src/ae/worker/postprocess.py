"""Post-processing: JSON and Excel output generation."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ae.config import get_settings
from ae.models import Document, Extraction, SchemaVersion, Task

logger = logging.getLogger(__name__)


def export_json(
    session: Session,
    task: Task,
    output_path: Path | None = None,
) -> Path:
    """Export extraction results to JSON."""
    settings = get_settings()
    if output_path is None:
        output_path = settings.output_path / f"{task.name}_results.json"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    extractions = _get_latest_extractions(session, task)

    results = []
    for ext, doc in extractions:
        entry = {
            "document": doc.filename,
            "file_hash": doc.file_hash,
            "extraction": ext.result or {},
            "confidence": ext.field_confidences or {},
            "overall_confidence": ext.overall_confidence,
            "iteration": ext.iteration,
            "status": ext.status,
        }
        if doc.metadata_extracted:
            entry["filename_metadata"] = doc.metadata_extracted
        results.append(entry)

    output = {
        "task": task.name,
        "description": task.description,
        "iteration": task.iteration,
        "total_documents": len(results),
        "results": results,
    }

    output_path.write_text(
        json.dumps(output, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    logger.info("Exported JSON to %s", output_path)
    return output_path


def export_excel(
    session: Session,
    task: Task,
    output_path: Path | None = None,
) -> Path:
    """Export extraction results to Excel."""
    settings = get_settings()
    if output_path is None:
        output_path = settings.output_path / f"{task.name}_results.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Get schema
    schema_ver = (
        session.query(SchemaVersion)
        .filter_by(task_id=task.id, is_active=True)
        .first()
    )
    if not schema_ver:
        raise ValueError("No active schema found")

    fields = schema_ver.schema_def.get("fields", [])
    field_names = [f.get("name", "") for f in fields]

    extractions = _get_latest_extractions(session, task)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction Results"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write headers
    headers = ["Document", "Status", "Confidence"] + field_names
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, (ext, doc) in enumerate(extractions, 2):
        ws.cell(row=row_idx, column=1, value=doc.filename)
        ws.cell(row=row_idx, column=2, value=ext.status)
        ws.cell(row=row_idx, column=3, value=round(ext.overall_confidence or 0, 3))

        result = ext.result or {}
        for col_offset, field_name in enumerate(field_names):
            value = result.get(field_name)
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            ws.cell(row=row_idx, column=4 + col_offset, value=value)

    # Confidence sheet
    ws2 = wb.create_sheet("Confidence Scores")
    conf_headers = ["Document"] + field_names
    for col, header in enumerate(conf_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, (ext, doc) in enumerate(extractions, 2):
        ws2.cell(row=row_idx, column=1, value=doc.filename)
        confidences = ext.field_confidences or {}
        for col_offset, field_name in enumerate(field_names):
            conf = confidences.get(field_name)
            cell = ws2.cell(row=row_idx, column=2 + col_offset, value=round(conf, 3) if conf else None)

            # Color code: red < 0.5, yellow 0.5-0.8, green > 0.8
            if conf is not None:
                if conf < 0.5:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif conf < 0.8:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Auto-width
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = 0
            for cell in col:
                try:
                    val = str(cell.value or "")
                    max_len = max(max_len, min(len(val), 50))
                except Exception:
                    pass
            ws_sheet.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(str(output_path))
    logger.info("Exported Excel to %s", output_path)
    return output_path


def _get_latest_extractions(
    session: Session,
    task: Task,
) -> list[tuple[Extraction, Document]]:
    """Get the latest extraction for each document in a task."""
    from sqlalchemy import func

    # Subquery: max extraction id per document
    subq = (
        session.query(
            Extraction.document_id,
            func.max(Extraction.id).label("max_id"),
        )
        .join(Document)
        .filter(Document.task_id == task.id)
        .group_by(Extraction.document_id)
        .subquery()
    )

    results = (
        session.query(Extraction, Document)
        .join(Document, Extraction.document_id == Document.id)
        .join(subq, Extraction.id == subq.c.max_id)
        .order_by(Document.filename)
        .all()
    )

    return results
